from openpyxl import load_workbook
import re


def extract_e1(path):
    """Extracts the Excel file to a matrix representation and primes it for interpretation"""
    wb = load_workbook(path, data_only=True)

    helmet_type = re.search(r"(CYCLING|SNOW)", path)[0]
    stage_of_testing = re.search(r"(Development|BatchTesting|Certificates)", path)[0]

    model = ""
    model_match = re.search(r"(CYCLING|SNOW)[\\/][\w\s+°]+", path)
    if model_match:
        model = re.sub(r"(CYCLING|SNOW)[\\/]", "", model_match[0])
    data = {"workbook": [], "helmet_type": helmet_type, "model": model, "stage_of_testing": stage_of_testing}

    for sheet in wb:
        # for now only handle table sheets
        if "Hit Locations" in sheet.title:
            continue

        # print("Work sheet: {}\n".format(sheet.title))

        rows = []
        r_max = sheet.max_row

        # add all the rows to the matrix
        for row in sheet.iter_rows(min_row=0, max_col=16, max_row=r_max, values_only=True):
            rows.append(row)

        data["workbook"].append(rows)

    print("num rows (E1): {}".format(len(rows)))

    return data
