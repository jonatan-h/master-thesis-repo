import zipfile

from openpyxl import load_workbook
import xlrd
from modules.extraction.PDFPlumber import extract_raw_text
import re
import pandas as pd


def is_pdf1(raw_text):
    for page in raw_text:
        if re.search(r"Impact(\s*(Uni|Tri)-Axial)", page):
            return True

    return False


def is_pdf2(raw_text):
    return re.search(r"Certification Institute 1", raw_text[0])


def is_pdf3(raw_text):
    return re.search(r"Certification Instittute 2", raw_text[0])


def is_pdf4(raw_text):
    return raw_text[0].startswith("Certification Institute 3")


def is_pdf5(raw_text):
    return re.search("Certification Institute 4", raw_text[0])


def get_active_sheet(workbook):
    active_sheet = 0
    index = 0
    for sheet in workbook.sheets():
        if sheet.sheet_selected:
            active_sheet = index
            break
        index += 1
    return active_sheet


def report_type(path):
    # TODO: Change the type format to be using strings, e.g. "E2" (or more intuitive numbered formats)
    filetype = str(path).split(".").pop()
    # currently only testing for "E1" type reports, rest rejected
    if filetype == "xlsx":
        print("File type: xlsx")

        try:
            wb = load_workbook(path, data_only=True)
        except zipfile.BadZipfile:
            print("failed to open workbook: \n")
            return -1

        ws = wb.active
        c = ws['A1']

        # TODO: add support for reports where the table is not in active sheet
        # TODO: add support for reports that don't have the row with the license in the table (E1 type)
        if c.value:
            if re.match(r"D[Aa][Tt][Ee]", str(c.value)):
                # print("Report Type: E1")
                return "E1"

        print("Report Type: other")
        return -1  # -1 for now means that there is no support for this type of report
    elif filetype == "xls":
        try:
            wb = xlrd.open_workbook(path)

            ws = wb.sheet_by_index(get_active_sheet(wb))
            first_row = ws.cell_value(rowx=0, colx=0)
            # print("A1: {0}".format(first_row))

            print("File type: xls")

            if first_row == '东 莞 益 安 运 动 用 品 有 限 公 司':
                # print("Report Type: E2")
                return "E2"
            return -1
        except (xlrd.biffh.XLRDError, AssertionError):

            print("failed to open workbook\n")
            return -1

    elif filetype.lower() == "pdf":
        raw_text = extract_raw_text(path)
        if is_pdf1(raw_text):
            return "PDF1"
        elif is_pdf2(raw_text):
            return "PDF2"
        elif is_pdf3(raw_text):
            return "PDF3"
        elif is_pdf4(raw_text):
            return "PDF4"
        elif is_pdf5(raw_text):
            return "PDF5"
        # TODO: extend further once more pdfs are added
        else:
            print("Report type not supported\n")
            return -1

    print("File type: not Excel or PDF\n")
    return -1  # -1 for now means that there is no support for this type of report
